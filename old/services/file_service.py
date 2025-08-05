# ファイル処理サービス
import os
import magic
import aiofiles
from typing import List, Dict, Any, Optional, BinaryIO
from fastapi import UploadFile, HTTPException
from pathlib import Path
import hashlib
from datetime import datetime
import logging

# ドキュメント処理用のインポート
from docx import Document as DocxDocument
import PyPDF2
import openpyxl
from PIL import Image
import csv
import json

logger = logging.getLogger(__name__)

class FileService:
    """ファイル処理関連のビジネスロジック"""
    
    def __init__(self):
        # 設定値
        self.max_file_size = int(os.getenv("MAX_FILE_SIZE", "52428800"))  # 50MB
        self.allowed_extensions = os.getenv("ALLOWED_FILE_EXTENSIONS", 
                                          "pdf,docx,pptx,xlsx,csv,txt,md,png,jpg,gif").split(",")
        self.upload_dir = Path("uploads")
        self.upload_dir.mkdir(exist_ok=True)
        
        # MIME型とファイル拡張子のマッピング
        self.mime_mapping = {
            "application/pdf": "pdf",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "docx",
            "application/vnd.openxmlformats-officedocument.presentationml.presentation": "pptx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
            "text/csv": "csv",
            "text/plain": "txt",
            "text/markdown": "md",
            "image/png": "png",
            "image/jpeg": "jpg",
            "image/gif": "gif"
        }
    
    async def validate_file(self, file: UploadFile) -> Dict[str, Any]:
        """ファイルのバリデーション"""
        try:
            # ファイルサイズチェック
            file_content = await file.read()
            file_size = len(file_content)
            
            if file_size > self.max_file_size:
                return {
                    "valid": False, 
                    "error": f"ファイルサイズが制限を超えています（制限: {self.max_file_size // 1024 // 1024}MB）"
                }
            
            if file_size == 0:
                return {"valid": False, "error": "空のファイルはアップロードできません"}
            
            # ファイル拡張子チェック
            file_extension = Path(file.filename).suffix.lower().lstrip('.')
            if file_extension not in self.allowed_extensions:
                return {
                    "valid": False, 
                    "error": f"許可されていないファイル形式です（許可形式: {', '.join(self.allowed_extensions)}）"
                }
            
            # MIME型チェック
            mime_type = magic.from_buffer(file_content, mime=True)
            if mime_type not in self.mime_mapping:
                return {"valid": False, "error": "不正なファイル形式です"}
            
            # 拡張子とMIME型の整合性チェック
            expected_extension = self.mime_mapping[mime_type]
            if file_extension != expected_extension:
                return {"valid": False, "error": "ファイル拡張子とファイル内容が一致しません"}
            
            # ファイルをリセット
            await file.seek(0)
            
            return {
                "valid": True,
                "file_size": file_size,
                "mime_type": mime_type,
                "extension": file_extension
            }
            
        except Exception as e:
            logger.error(f"ファイル検証エラー: {e}")
            return {"valid": False, "error": "ファイル検証中にエラーが発生しました"}
    
    async def save_file(self, file: UploadFile, project_id: int, user_id: int) -> Dict[str, Any]:
        """ファイルを保存"""
        try:
            # ファイルバリデーション
            validation_result = await self.validate_file(file)
            if not validation_result["valid"]:
                return {"success": False, "message": validation_result["error"]}
            
            # ファイル名の生成（重複回避のためハッシュを含める）
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_hash = hashlib.md5(f"{user_id}_{project_id}_{file.filename}_{timestamp}".encode()).hexdigest()[:8]
            safe_filename = f"{timestamp}_{file_hash}_{file.filename}"
            
            # 保存パスの作成
            project_dir = self.upload_dir / f"project_{project_id}"
            project_dir.mkdir(exist_ok=True)
            file_path = project_dir / safe_filename
            
            # ファイルを保存
            async with aiofiles.open(file_path, 'wb') as f:
                content = await file.read()
                await f.write(content)
            
            logger.info(f"ファイル保存成功: {file_path}")
            
            return {
                "success": True,
                "file_path": str(file_path),
                "file_name": file.filename,
                "file_size": validation_result["file_size"],
                "file_type": validation_result["extension"],
                "mime_type": validation_result["mime_type"]
            }
            
        except Exception as e:
            logger.error(f"ファイル保存エラー: {e}")
            return {"success": False, "message": f"ファイル保存に失敗しました: {str(e)}"}
    
    async def delete_file(self, file_path: str) -> bool:
        """ファイルを削除"""
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
                logger.info(f"ファイル削除成功: {file_path}")
                return True
            return False
        except Exception as e:
            logger.error(f"ファイル削除エラー: {e}")
            return False
    
    async def extract_text_from_file(self, file_path: str, file_type: str) -> str:
        """ファイルからテキストを抽出"""
        try:
            if file_type == "pdf":
                return await self._extract_from_pdf(file_path)
            elif file_type == "docx":
                return await self._extract_from_docx(file_path)
            elif file_type == "xlsx":
                return await self._extract_from_xlsx(file_path)
            elif file_type == "csv":
                return await self._extract_from_csv(file_path)
            elif file_type in ["txt", "md"]:
                return await self._extract_from_text(file_path)
            elif file_type in ["png", "jpg", "gif"]:
                return await self._extract_from_image(file_path)
            else:
                return ""
                
        except Exception as e:
            logger.error(f"テキスト抽出エラー ({file_type}): {e}")
            return ""
    
    async def _extract_from_pdf(self, file_path: str) -> str:
        """PDFからテキストを抽出"""
        text = ""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
        except Exception as e:
            logger.error(f"PDF抽出エラー: {e}")
        return text
    
    async def _extract_from_docx(self, file_path: str) -> str:
        """Wordドキュメントからテキストを抽出"""
        text = ""
        try:
            doc = DocxDocument(file_path)
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
        except Exception as e:
            logger.error(f"DOCX抽出エラー: {e}")
        return text
    
    async def _extract_from_xlsx(self, file_path: str) -> str:
        """Excelファイルからテキストを抽出"""
        text = ""
        try:
            workbook = openpyxl.load_workbook(file_path)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"[シート: {sheet_name}]\n"
                for row in sheet.iter_rows():
                    row_text = ""
                    for cell in row:
                        if cell.value:
                            row_text += str(cell.value) + "\t"
                    if row_text.strip():
                        text += row_text + "\n"
                text += "\n"
        except Exception as e:
            logger.error(f"XLSX抽出エラー: {e}")
        return text
    
    async def _extract_from_csv(self, file_path: str) -> str:
        """CSVファイルからテキストを抽出"""
        text = ""
        try:
            with open(file_path, 'r', encoding='utf-8', newline='') as file:
                csv_reader = csv.reader(file)
                for row in csv_reader:
                    text += "\t".join(row) + "\n"
        except UnicodeDecodeError:
            try:
                with open(file_path, 'r', encoding='shift_jis', newline='') as file:
                    csv_reader = csv.reader(file)
                    for row in csv_reader:
                        text += "\t".join(row) + "\n"
            except Exception as e:
                logger.error(f"CSV抽出エラー: {e}")
        except Exception as e:
            logger.error(f"CSV抽出エラー: {e}")
        return text
    
    async def _extract_from_text(self, file_path: str) -> str:
        """テキストファイルからテキストを抽出"""
        text = ""
        try:
            async with aiofiles.open(file_path, 'r', encoding='utf-8') as file:
                text = await file.read()
        except UnicodeDecodeError:
            try:
                async with aiofiles.open(file_path, 'r', encoding='shift_jis') as file:
                    text = await file.read()
            except Exception as e:
                logger.error(f"テキスト抽出エラー: {e}")
        except Exception as e:
            logger.error(f"テキスト抽出エラー: {e}")
        return text
    
    async def _extract_from_image(self, file_path: str) -> str:
        """画像ファイルからメタデータを抽出（OCRは今後実装）"""
        text = ""
        try:
            with Image.open(file_path) as img:
                # 基本的なメタデータを抽出
                text += f"画像サイズ: {img.size[0]}x{img.size[1]}\n"
                text += f"画像モード: {img.mode}\n"
                text += f"ファイル形式: {img.format}\n"
                
                # EXIFデータがあれば抽出
                if hasattr(img, '_getexif') and img._getexif():
                    exif_data = img._getexif()
                    text += "EXIF情報:\n"
                    for tag, value in exif_data.items():
                        text += f"  {tag}: {value}\n"
                
                # 将来的にはOCRを実装予定
                text += "\n[注意: 画像内のテキスト認識は未実装です]\n"
                
        except Exception as e:
            logger.error(f"画像メタデータ抽出エラー: {e}")
        return text
    
    async def get_file_stats(self, project_id: Optional[int] = None) -> Dict[str, Any]:
        """ファイル統計情報を取得"""
        try:
            if project_id:
                project_dir = self.upload_dir / f"project_{project_id}"
                if not project_dir.exists():
                    return {"total_files": 0, "total_size": 0}
                
                files = list(project_dir.glob("*"))
            else:
                files = list(self.upload_dir.rglob("*"))
            
            total_files = len([f for f in files if f.is_file()])
            total_size = sum(f.stat().st_size for f in files if f.is_file())
            
            return {
                "total_files": total_files,
                "total_size": total_size,
                "total_size_mb": round(total_size / 1024 / 1024, 2)
            }
            
        except Exception as e:
            logger.error(f"ファイル統計取得エラー: {e}")
            return {"total_files": 0, "total_size": 0, "total_size_mb": 0}